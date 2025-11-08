# -*- coding: utf-8 -*-
"""
Testes automatizados para o backend FastAPI do Diário Operacional

Para rodar:
  pip install pytest pytest-asyncio httpx --break-system-packages
  pytest test_app.py -v

Para rodar com coverage:
  pip install pytest-cov --break-system-packages
  pytest test_app.py --cov=app --cov-report=html
"""

import pytest
import io
import os
from datetime import datetime
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import Workbook

# Configurar variável de ambiente para testes
os.environ["SUPABASE_DB_URL"] = "sqlite:///./test.db"  # Usar SQLite para testes

# Importar a aplicação DEPOIS de configurar o env
from app import app, engine


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture(scope="module")
def client():
    """Cliente de teste HTTP"""
    return TestClient(app)


@pytest.fixture(scope="function", autouse=True)
def setup_database():
    """Criar e limpar banco de testes antes de cada teste"""
    # Criar tabelas
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS uploads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client TEXT NOT NULL,
                ym TEXT NOT NULL,
                kind TEXT NOT NULL,
                data BLOB NOT NULL,
                hash TEXT,
                created_at TEXT NOT NULL
            )
        """))
    
    yield
    
    # Limpar banco após o teste
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM uploads"))


@pytest.fixture
def sample_booking_excel():
    """Gerar Excel de booking de exemplo"""
    wb = Workbook()
    ws = wb.active
    
    # Cabeçalhos
    ws.append(["DATA_BOOKING", "NOME_FANTASIA", "QTDE_CONTAINER", "BOOKING", 
               "SIGLA_PORTO_ORIGEM", "SIGLA_PORTO_DESTINO", "DESC_STATUS"])
    
    # Dados
    ws.append(["2024-10-15", "Cliente A", 10, "BKG123", "SANTOS", "BUENOS AIRES", "Ativo"])
    ws.append(["2024-10-20", "Cliente B", 5, "BKG124", "PARANAGUÁ", "MONTEVIDEO", "Ativo"])
    ws.append(["2024-11-05", "Cliente A", 8, "BKG125", "SANTOS", "BUENOS AIRES", "Ativo"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def sample_multimodal_excel():
    """Gerar Excel de multimodal de exemplo"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Cliente", "Porto da Operação", "Tipo de Operação", "Agendamento", 
               "Causador Reagenda", "Área Responsável", "Justificativa Reagendamento"])
    ws.append(["Cliente A", "SANTOS", "Importação", "2024-10-15", "EMBARCADOR", 
               "OPERACIONAL", "Atraso na documentação"])
    ws.append(["Cliente B", "PARANAGUÁ", "Exportação", "2024-10-20", "", "", ""])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def sample_transportes_excel():
    """Gerar Excel de transportes de exemplo"""
    wb = Workbook()
    ws = wb.active
    
    ws.append(["Embarcador", "Data Coleta", "Data Entrega", "Porto", "Tipo Operação"])
    ws.append(["Cliente A", "2024-10-10", "2024-10-15", "SANTOS", "FCL"])
    ws.append(["Cliente B", "2024-10-18", "2024-10-22", "PARANAGUÁ", "LCL"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# TESTES - HEALTH CHECK
# =============================================================================

def test_health_check(client):
    """Teste: Endpoint de health check"""
    response = client.get("/api/health")
    assert response.status_code == 200
    data = response.json()
    assert data["ok"] is True
    assert "cache_size" in data


# =============================================================================
# TESTES - UPLOAD
# =============================================================================

def test_upload_success(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Upload de arquivos com sucesso"""
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    response = client.post(
        "/api/upload",
        files=files,
        data={"client": "TEST_CLIENT"}
    )
    
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "ok"
    assert "periods" in data
    assert len(data["periods"]) > 0
    assert "embarcadores" in data
    assert len(data["embarcadores"]) > 0


def test_upload_missing_file(client, sample_booking_excel):
    """Teste: Upload falhando por falta de arquivo"""
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        # Faltando multimodal e transportes
    }
    
    response = client.post(
        "/api/upload",
        files=files,
        data={"client": "TEST_CLIENT"}
    )
    
    # Deve retornar erro 422 (validation error)
    assert response.status_code == 422


def test_upload_deduplication(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Deduplicação de uploads idênticos"""
    files = {
        "booking": ("booking.xlsx", sample_booking_excel.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    # Primeiro upload
    response1 = client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    assert response1.status_code == 200
    data1 = response1.json()
    assert len(data1["inserted"]) > 0
    
    # Segundo upload (idêntico)
    sample_booking_excel.seek(0)
    sample_multimodal_excel.seek(0)
    sample_transportes_excel.seek(0)
    
    files2 = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    response2 = client.post("/api/upload", files=files2, data={"client": "TEST_CLIENT"})
    assert response2.status_code == 200
    data2 = response2.json()
    # Deve ter skipped por hash igual
    assert len(data2.get("skipped", [])) > 0


# =============================================================================
# TESTES - AVAILABLE DATA
# =============================================================================

def test_available_data_empty(client):
    """Teste: Buscar dados disponíveis quando não há dados"""
    response = client.get("/api/available-data?client=EMPTY_CLIENT")
    assert response.status_code == 200
    data = response.json()
    assert data["has_data"] is False
    assert len(data["periods"]) == 0
    assert len(data["embarcadores"]) == 0


def test_available_data_with_uploads(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Buscar dados disponíveis após upload"""
    # Fazer upload primeiro
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    upload_response = client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    assert upload_response.status_code == 200
    
    # Buscar dados disponíveis
    response = client.get("/api/available-data?client=TEST_CLIENT")
    assert response.status_code == 200
    data = response.json()
    assert data["has_data"] is True
    assert len(data["periods"]) > 0
    assert len(data["embarcadores"]) > 0


# =============================================================================
# TESTES - SUMMARY
# =============================================================================

def test_summary_without_data(client):
    """Teste: Buscar resumo sem dados no banco"""
    response = client.get("/api/summary?client=TEST_CLIENT&ym=2024-10&embarcador=Cliente A")
    assert response.status_code == 400  # Falta de dados


def test_summary_with_data(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Buscar resumo com dados"""
    # Upload
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    
    # Buscar resumo
    response = client.get("/api/summary?client=TEST_CLIENT&ym=2024-10&embarcador=Cliente A")
    assert response.status_code == 200
    data = response.json()
    assert "kpis" in data
    assert "debug" in data


# =============================================================================
# TESTES - FLUSH
# =============================================================================

def test_flush_all(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Limpar todos os dados de um cliente"""
    # Upload
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    
    # Flush
    response = client.delete("/api/flush?client=TEST_CLIENT")
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "ok"
    assert data["deleted"] > 0


def test_flush_specific_period(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Limpar dados de um período específico"""
    # Upload
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    
    # Flush de período específico
    response = client.delete("/api/flush?client=TEST_CLIENT&ym=2024-10")
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "ok"
    assert data["detail"]["ym"] == "2024-10"


# =============================================================================
# TESTES - CACHE
# =============================================================================

def test_clear_cache(client):
    """Teste: Limpar cache"""
    response = client.get("/api/clear-cache")
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "ok"
    assert "message" in data


# =============================================================================
# TESTES - GERAÇÃO DE EMAIL (OPCIONAL - REQUER GEMINI)
# =============================================================================

@pytest.mark.skipif(not os.getenv("GEMINI_API_KEY"), reason="GEMINI_API_KEY não configurada")
def test_generate_email(client, sample_booking_excel, sample_multimodal_excel, sample_transportes_excel):
    """Teste: Gerar email com IA (requer GEMINI_API_KEY)"""
    # Upload
    files = {
        "booking": ("booking.xlsx", sample_booking_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "multimodal": ("multi.xlsx", sample_multimodal_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "transportes": ("transp.xlsx", sample_transportes_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }
    
    client.post("/api/upload", files=files, data={"client": "TEST_CLIENT"})
    
    # Gerar email
    payload = {
        "client": "TEST_CLIENT",
        "yms": ["2024-10"],
        "embarcadores": ["Cliente A"]
    }
    
    response = client.post("/api/generate-email", json=payload)
    assert response.status_code == 200
    data = response.json()
    assert "email" in data
    assert "email_html" in data


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])